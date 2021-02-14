from datetime import datetime
from openpyxl import Workbook 
from openpyxl import load_workbook
import os
from selenium import webdriver
import requests
from webdriver_manager.chrome import ChromeDriverManager
import urllib.request
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep

driver = webdriver.Chrome(ChromeDriverManager().install())

username = 'your_uesrname'
password = 'your_password'

def get_verb(verb):
    driver.get('https://www.verbix.com/webverbix/German/{}.html'.format(verb))
    past = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/section/div/div[3]/div[1]/div[3]/table/tbody/tr[1]/td[2]')
    perfect = driver.find_element_by_xpath('//*[@id="verbixConjugations"]/section/div/div[3]/div[1]/div[2]/table/tbody/tr[3]/td[2]')
    return past.text+' ;'+perfect.text

def get_sound(word):
    url = "https://www.howtopronounce.com/german/{}".format(word)
    driver.get(url)
    url2 = driver.find_element_by_xpath("/html/body/main/div/section[2]/div[1]/div[2]/div/section[2]/div/article[1]/div/div/span/audio").get_attribute('src')
    re = requests.get(url2)
    with open('words/tmp/{}.mp3'.format(word),'wb') as f:
        f.write(re.content)


def memrise(data,words):
    driver.get("https://app.memrise.com/signin")
    username = driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/form/div[1]/div/input')
    username.send_keys(username)
    password = driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/form/div[2]/div/input')
    password.send_keys(password)
    button = driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/form/div[3]/div[1]/button')
    button.click()
    sleep(5)
    driver.get('https://app.memrise.com/course/5947047/german-b1-word/edit')
    button = driver.find_element_by_xpath('/html/body/div[3]/div[5]/div/div[1]/div[2]/div/button')
    button.click()
    button = driver.find_element_by_xpath('/html/body/div[3]/div[5]/div/div[1]/div[2]/div/ul/li[1]/a')
    button.click()
    sleep(6)
    levels = driver.find_elements_by_class_name('level')
    current = levels[-1]
    actions = ActionChains(driver)
    actions.move_to_element(current).perform()

    button = driver.find_element_by_xpath('/html/body/div[3]/div[5]/div/div[2]/div[{}]/div[2]/div/div/div/button'.format(len(levels)))
    # print(button.text)
    button.click()
    button = driver.find_element_by_xpath('/html/body/div[3]/div[5]/div/div[2]/div[{}]/div[2]/div/div/div/ul/li/a'.format(len(levels)))
    button.click()
    sleep(3)
    # inp = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/form/div[2]')
    # inp.send_keys(data)
    pop = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/form/div[2]/textarea')
    pop.send_keys(data)
    sleep(1)
    button = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[2]/form/div[1]/div/label[2]/input')
    button.click()
    button = driver.find_element_by_xpath('/html/body/div[4]/div/div/div[3]/a[2]')
    button.click()
    sleep(5)
    for i in range(len(words)):
        audio = driver.find_element_by_xpath('/html/body/div[3]/div[5]/div/div[2]/div[{level}]/div[3]/table/tbody[1]/tr[{word}]/td[5]/div/div[1]/input'.format(level = len(levels),word=i+1))
        #actions.move_to_element(audio).perform()
        sleep(2)
        wd = words[i]
        if words[i] == 'skip':
            continue
        else:
            audio.send_keys('/home/mjandar12/Desktop/german/words/tmp/{}.mp3'.format(words[i]))










def store_words(german, meaning):
    wb = Workbook()
    ws = wb.active
    d = os.listdir('words')
    artikel = ['der','das','die','Der','Das','Die']
    words = []
    words_m = []
    verbs = []
    verbs_m = []
    verbs_past = []
    adjectives = []
    adjectives_m = []
    data = ''
    for c,word in enumerate(german):
        if word[:3] in artikel:
            german[c] = word[4:]
            words_m.append(meaning[c])
            words.append(word)
            data += '{word},{meaning}\n'.format(word = word,meaning=meaning[c])
        elif word[-2:] == 'en':
            verbs_m.append(meaning[c])
            verbs.append(word)
            verbs_past.append(get_verb(word))
            data += '{word},{meaning},{extra}\n'.format(word = word,meaning=meaning[c],extra = verbs_past[-1])
        else:
            adjectives_m.append(meaning[c])
            adjectives.append(word)
            data += '{word},{meaning}\n'.format(word = word,meaning=meaning[c])
        if word[-1] != '-':
            get_sound(german[c])

        
        
        

    ws.cell(row=1,column=1).value = "nouns"
    n = 2
    for i in range(len(words)):
        ws.cell(row=n, column=1).value = words[i]
        ws.cell(row=n, column=2).value = words_m[i]
        n +=1
    ws.cell(row=n,column=1).value = "verbs"
    n +=1
    for i in range(len(verbs)):
        ws.cell(row=n, column=1).value = verbs[i]
        ws.cell(row=n, column=2).value = verbs_m[i]
        ws.cell(row=n, column=3).value = verbs_past[i]
        n +=1
    ws.cell(row=n,column=1).value = "adjectives"
    n +=1
    for i in range(len(adjectives)):
        ws.cell(row=n, column=1).value = adjectives[i]
        ws.cell(row=n, column=2).value = adjectives_m[i]
        n +=1
    wb.save("words/group{}.xlsx".format(len(d)))
    memrise(data,german)





def word_input():
    words = []
    meanings = []

    while True:
        word = input("german word: ")
        if word == 'save -a':
            store_words(words,meanings)
            memrise(words,meanings)
            break
        elif word == 'quit':
            store_words(words,meanings)
            break
        meaning = input("meaning: ")
        words.append(word)
        meanings.append(meaning)









inp = input("number of words to enter: ")
if inp == "i":
    word_input()
elif inp == "load":
    file = int(input("file number to load:"))
    wb = load_workbook("words/group{}.xlsx".format(file))
    sheet = wb.active
    n = 2
    st = 0
    words = []
    data = ''
    while True:
        
        cell1 = sheet.cell(row=n,column=1).value
        if cell1 == None:
            break
        if cell1 == "verbs":
            st = 1
        elif cell1 == "adjectives":
            st = 2
        else:
            word = cell1
            meaning = sheet.cell(row=n,column=2).value
            if st ==1 :
                data = data + word + "," + meaning + '\n'
            else:
                data = data + word + "," + meaning +'\n'
            if word[-1] != '-':
                if st == 0:
                    get_sound(word[4:])
                    words.append(word[4:])
                else:
                    get_sound(word)
                    words.append(word)
            else:
                words.append("skip")
            
            print(data)
        n+=1
    print(data)
    memrise(data,words)







driver.close()
